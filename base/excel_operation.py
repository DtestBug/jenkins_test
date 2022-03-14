from openpyxl import load_workbook
import configparser


def r_base_conf(conf_path, tag, data):
    """
    :param conf_path: conf文件路径
    :param tag: [tag]
    :param data: [配置]
    :return: [配置内容]
    """
    obj = configparser.ConfigParser()
    obj.read(conf_path, encoding='utf-8')
    return obj[tag][data]


def read_excel(file_path, sheet_name):
    """
    :param file_path: Excel路径
    :param sheet_name: sheet名字
    :return:
    """
    # 读取Excel表格表面数据，不读取公式，如需获取表格内公式，将data_only改为false(本代码不推荐，因为改了之后不会显示Excel数据)
    file_sheet = load_workbook(file_path, data_only=True, )[sheet_name]

    # 获取头部
    sheet_header = []
    for Header_data in range(file_sheet.max_column):
        sheet_header.append(file_sheet.cell(row=1, column=Header_data + 1).value)

    # 获取数据
    data_list = []
    for data_data1 in range(file_sheet.max_row - 1):
        data_dict = {}
        for data_data2 in range(len(sheet_header)):
            data_dict[sheet_header[data_data2]] = file_sheet.cell(row=data_data1 + 2, column=data_data2 + 1).value
        data_list.append(data_dict)
    return data_list  # 返回数据


def read_sheets(file_path):
    """
    :param file_path: Excel路径
    :return: Excel中所有sheet
    """
    wb = load_workbook(file_path)
    res = wb.sheetnames
    return res


def rewrite(file_name, sheet_name, sheet_row, actual_result, test_result, runtime):
    """
    :param file_name: 文件路径+名字
    :param sheet_name: 表单名字
    :param sheet_row: 一般从0开始，需要放到循环外例如：x = 0
    :param actual_result: 预期结果
    :param test_result: 实际测试结果
    :param runtime: 运行时间
    :return: 以上数据追加进表
    """
    __rewrite = load_workbook(file_name)  # 追加数据表
    ws1 = __rewrite[sheet_name.title()]

    # .fill为颜色值 / .value为参数值
    ws1.cell(row=sheet_row + 1, column=9).value = actual_result  # 实际结果
    ws1.cell(row=sheet_row + 1, column=10).value = test_result  # Yes/No是否通过
    ws1.cell(row=sheet_row + 1, column=11).value = runtime  # 执行时间
    __rewrite.save(file_name)  # 保存追加数据后的表


def excel_data(if_sheet):
    """
    :param
        if classname == sheet_name
    :return
        sheet_data / errors
    """
    excel_path = r_base_conf("路径", "file", "excel_path")
    sheets = read_sheets(excel_path)
    for sheet in sheets:
        # if if_sheet in sheet:
        #     return read_excel(excel_path, sheet)
        # else:
        #     pass
        return read_excel(excel_path, sheet) if (if_sheet in sheet) else ""
        # return res
